"""import pandas as pd
import matplotlib.pyplot as plt

# CSV dosyasını oku
df = pd.read_csv("../../csharp/SensorSimulator/sensor_data.csv", encoding="utf-8")

# İlk 5 satırı göster
print("İlk 5 veri satırı:")
print(df.head())

# Temel istatistikler
print("\nİstatistikler:")
print("Ortalama sıcaklık:", df["Sıcaklık"].mean())
print("Maksimum sıcaklık:", df["Sıcaklık"].max())
print("Minimum sıcaklık:", df["Sıcaklık"].min())

# Zaman serisi grafiği
plt.figure(figsize=(10,5))
plt.plot(df["Tarih"], df["Sıcaklık"], marker="o")
plt.xticks(rotation=45)
plt.xlabel("Zaman")
plt.ylabel("Sıcaklık (°C)")
plt.title("Zaman İçinde Sıcaklık Değişimi")
plt.tight_layout()
plt.show()
GÜN 3
"""
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Veriyi oku
df = pd.read_csv("../../csharp/SensorSimulator/sensor_data.csv", usecols=[0,1])
df = df.tail(100)  # sadece son 200 satır

# Tarihi datetime formatına çevir
df["Tarih"] = pd.to_datetime(df["Tarih"])

# Basit istatistikler
"""print("Ortalama:", df["Sıcaklık"].mean())
print("Maksimum:", df["Sıcaklık"].max())
print("Minimum:", df["Sıcaklık"].min()) GÜN 4"""

ortalama = df["Sıcaklık"].mean()
maksimum = df["Sıcaklık"].max()
minimum = df["Sıcaklık"].min()

print("Ortalama:", ortalama)
print("Maksimum:", maksimum)
print("Minimum:", minimum)


# Grafik çiz
plt.figure(figsize=(10,5))
plt.plot(df["Tarih"], df["Sıcaklık"], label="Sıcaklık", color="blue")
plt.xlabel("Zaman")
plt.ylabel("Sıcaklık (°C)")
plt.title("Zaman İçinde Sıcaklık Değişimi")
plt.legend()
plt.grid()

# Grafiği kaydet
"""plt.savefig("temperature_plot.png")

print("Grafik 'temperature_plot.png' dosyasına kaydedildi.") GÜN 4"""

plot_filename = "temperature_plot.png"
plt.savefig(plot_filename)
plt.close()

# Excel dosyası oluştur
wb = Workbook()
ws = wb.active
ws.title = "Analiz Raporu"

# İstatistikleri yaz
ws["A1"] = "Ortalama"
ws["B1"] = ortalama
ws["A2"] = "Maksimum"
ws["B2"] = maksimum
ws["A3"] = "Minimum"
ws["B3"] = minimum

# Grafiği ekle
img = Image(plot_filename)
ws.add_image(img, "D1")

# Kaydet
wb.save("analysis_report.xlsx")

print("Rapor 'analysis_report.xlsx' dosyasına kaydedildi.")

# --- ANOMALİ TESPİTİ ---
import numpy as np

"""# Ortalama ve standart sapma
mean = df["Sıcaklık"].mean()
std = df["Sıcaklık"].std()

# Anomaliyi tanımla (ortalamanın ±2*std dışındaki değerler)
df["Anomali"] = np.where((df["Sıcaklık"] > mean + 2*std) | (df["Sıcaklık"] < mean - 2*std), True, False)

print("\nTespit edilen anomaliler:")
print(df[df["Anomali"] == True])

# Anomaliyi grafikte göster
plt.figure(figsize=(10,6))
plt.plot(df["Tarih"], df["Sıcaklık"], label="Sıcaklık", color="blue")
plt.scatter(df[df["Anomali"] == True]["Tarih"], df[df["Anomali"] == True]["Sıcaklık"], color="red", label="Anomali")
plt.xlabel("Zaman")
plt.ylabel("Sıcaklık (°C)")
plt.title("Sıcaklık Verileri ve Anomaliler")
plt.legend()
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("anomaly_plot.png")
plt.show()
"""


# Örnek veri: kendi df'ni kullanabilirsin
# df = pd.read_csv("veri.csv")

# Anomalileri tespit et (manuel sınırlar: <15 veya >35)
df["Anomali"] = (df["Sıcaklık"] < 15) | (df["Sıcaklık"] > 35)

print("\nTespit edilen anomaliler:")
print(df[df["Anomali"] == True])

# Grafikle göster
plt.figure(figsize=(10,6))
plt.plot(df["Tarih"], df["Sıcaklık"], label="Sıcaklık", color="blue")
plt.scatter(df[df["Anomali"] == True]["Tarih"],
            df[df["Anomali"] == True]["Sıcaklık"],
            color="red", label="Anomali")
plt.xlabel("Zaman")
plt.ylabel("Sıcaklık (°C)")
plt.title("Sıcaklık Verileri ve Anomaliler")
plt.legend()
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("anomaly_plot.png")
plt.show()

